"""Extract and enrich real ESCO + O*NET data into loader-ready files.

Sources
-------
ESCO v1.2.1 zip  — occupations_en.csv, skills_en.csv, occupationSkillRelations_en.csv
O*NET 2019 CSV   — 2019_Occupations.csv
O*NET 30.2 zip   — Occupation Data.xlsx, Job Zones.xlsx, Task Statements.xlsx,
                   Technology Skills.xlsx
O*NET txt files  — Skills.txt, Knowledge.txt, Technology Skills.txt,
                   Task Statements.txt, Education, Training, and Experience.txt
                   (download individually from onetcenter.org/database.html#individual-files)

Output (all written to --output-dir)
------
  esco_occupations_enriched.csv   3,043 occupations with essential+optional skills appended
  esco_skills.csv                 13,960 skill documents
  onet_occupations_enriched.csv   1,016 occupations with tasks + tech skills appended
  role_templates_real.json        ~900 role templates derived from O*NET 30.2 data

Usage
-----
  cd agents

  # ESCO + O*NET occupation CSVs only (no txt files needed):
  python -m scripts.prepare_real_data \\
    --esco-zip   "C:/Users/User/Downloads/ESCO dataset - v1.2.1 - classification - en - csv.zip" \\
    --onet-zip   "C:/Users/User/Downloads/db_30_2_excel.zip" \\
    --output-dir data/knowledge-base

  # Also build role templates (requires txt files):
  python -m scripts.prepare_real_data \\
    --esco-zip        "C:/Users/User/Downloads/ESCO dataset - v1.2.1 - classification - en - csv.zip" \\
    --onet-zip        "C:/Users/User/Downloads/db_30_2_excel.zip" \\
    --onet-txt-dir    "C:/Users/User/Downloads" \\
    --output-dir      data/knowledge-base \\
    --build-templates

  # Role templates only (skip ESCO + onet CSVs):
  python -m scripts.prepare_real_data \\
    --onet-zip        "C:/Users/User/Downloads/db_30_2_excel.zip" \\
    --onet-txt-dir    "C:/Users/User/Downloads" \\
    --output-dir      data/knowledge-base \\
    --build-templates --skip-esco --skip-onet
"""
from __future__ import annotations

import argparse
import csv
import io
import sys
import zipfile
import json
from collections import defaultdict
from pathlib import Path

_ONET_TASK_LIMIT = 12
_ONET_TECH_LIMIT = 15
_ESCO_ESSENTIAL_LIMIT = 20
_ESCO_OPTIONAL_LIMIT = 10


# ── ESCO ──────────────────────────────────────────────────────────────────────

def _read_esco_zip(zip_path: str) -> tuple[list[dict], list[dict], list[dict]]:
    """Return (occupations, skills, occ_skill_relations) as lists of dicts."""
    z = zipfile.ZipFile(zip_path)

    def read_csv(name: str) -> list[dict]:
        return list(csv.DictReader(
            io.StringIO(z.read(name).decode("utf-8-sig"))
        ))

    print("  Reading ESCO occupations_en.csv ...")
    occupations = read_csv("occupations_en.csv")
    print(f"    {len(occupations):,} occupations")

    print("  Reading ESCO skills_en.csv ...")
    skills = read_csv("skills_en.csv")
    print(f"    {len(skills):,} skills")

    print("  Reading ESCO occupationSkillRelations_en.csv ...")
    relations = read_csv("occupationSkillRelations_en.csv")
    print(f"    {len(relations):,} occupation-skill links")

    return occupations, skills, relations


def build_esco_occupations_enriched(
    occupations: list[dict],
    relations: list[dict],
) -> list[dict]:
    """Join each occupation with its essential and optional skills."""
    # Group skill labels by occupation URI and relation type
    essential: dict[str, list[str]] = defaultdict(list)
    optional: dict[str, list[str]] = defaultdict(list)

    for rel in relations:
        uri = rel.get("occupationUri", "")
        label = rel.get("skillLabel", "").strip()
        rel_type = rel.get("relationType", "").lower()
        if not uri or not label:
            continue
        if rel_type == "essential":
            essential[uri].append(label)
        else:
            optional[uri].append(label)

    rows: list[dict] = []
    for occ in occupations:
        uri = occ.get("conceptUri", "")
        label = occ.get("preferredLabel", "").strip()
        description = occ.get("description", "").strip()
        if not label or not description:
            continue

        ess = essential.get(uri, [])[:_ESCO_ESSENTIAL_LIMIT]
        opt = optional.get(uri, [])[:_ESCO_OPTIONAL_LIMIT]

        enriched_desc = description
        if ess:
            enriched_desc += "\n\nEssential skills:\n" + "\n".join(f"- {s}" for s in ess)
        if opt:
            enriched_desc += "\n\nOptional skills:\n" + "\n".join(f"- {s}" for s in opt)

        rows.append({
            "conceptUri": uri,
            "preferredLabel": label,
            "altLabels": occ.get("altLabels", "").replace("\n", ", "),
            "description": enriched_desc,
        })

    return rows


def build_esco_skills(skills: list[dict]) -> list[dict]:
    """Convert each ESCO skill into a document row."""
    rows: list[dict] = []
    for s in skills:
        label = s.get("preferredLabel", "").strip()
        description = s.get("description", "").strip()
        if not label or not description:
            continue

        uri = s.get("conceptUri", "")
        skill_type = s.get("skillType", "")
        reuse = s.get("reuseLevel", "")

        enriched = description
        if skill_type or reuse:
            enriched += f"\n\nType: {skill_type}. Reuse level: {reuse}."

        rows.append({
            "conceptUri": uri,
            "preferredLabel": label,
            "altLabels": s.get("altLabels", "").replace("\n", ", "),
            "description": enriched,
        })
    return rows


# ── O*NET ─────────────────────────────────────────────────────────────────────

def build_onet_from_csv(csv_path: str) -> list[dict]:
    """Build occupation rows from the O*NET 2019 standalone CSV."""
    rows: list[dict] = []
    with open(csv_path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            # Handle both 2019 column names and plain names
            code = (row.get("O*NET-SOC 2019 Code") or row.get("O*NET-SOC Code", "")).strip()
            title = (row.get("O*NET-SOC 2019 Title") or row.get("Title", "")).strip()
            description = (row.get("O*NET-SOC 2019 Description") or row.get("Description", "")).strip()
            if not title or not description:
                continue
            rows.append({
                "conceptUri": f"onet:{code}",
                "preferredLabel": title,
                "altLabels": "",
                "description": description,
            })
    return rows


def build_onet_enriched_from_zip(zip_path: str) -> list[dict]:
    """Build enriched O*NET 30.2 occupation docs from the Excel zip."""
    try:
        import openpyxl  # noqa: PLC0415
    except ImportError:
        print("  openpyxl not installed — skipping O*NET 30.2 Excel enrichment.", file=sys.stderr)
        return []

    z = zipfile.ZipFile(zip_path)

    def read_xlsx(name: str) -> list[tuple]:
        wb = openpyxl.load_workbook(
            io.BytesIO(z.read(f"db_30_2_excel/{name}")),
            read_only=True, data_only=True,
        )
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        return rows

    print("  Reading O*NET 30.2 Occupation Data.xlsx ...")
    occ_rows = read_xlsx("Occupation Data.xlsx")
    headers_occ = occ_rows[0]  # ('O*NET-SOC Code', 'Title', 'Description')
    occupations = {
        r[0]: {"title": r[1], "description": r[2]}
        for r in occ_rows[1:] if r[0] and r[1]
    }
    print(f"    {len(occupations):,} occupations")

    print("  Reading O*NET 30.2 Task Statements.xlsx ...")
    task_rows = read_xlsx("Task Statements.xlsx")
    # headers: O*NET-SOC Code, Title, Task ID, Task, Task Type, ...
    tasks: dict[str, list[str]] = defaultdict(list)
    for r in task_rows[1:]:
        code, title, task_id, task, task_type = r[0], r[1], r[2], r[3], r[4]
        if code and task and str(task_type).lower() == "core":
            tasks[code].append(str(task))
    print(f"    {sum(len(v) for v in tasks.values()):,} core tasks across {len(tasks):,} occupations")

    print("  Reading O*NET 30.2 Technology Skills.xlsx ...")
    tech_rows = read_xlsx("Technology Skills.xlsx")
    # headers: O*NET-SOC Code, Title, Example, Commodity Code, Commodity Title, Hot Technology, In Demand
    tech: dict[str, list[str]] = defaultdict(list)
    for r in tech_rows[1:]:
        code, example, commodity_title = r[0], r[2], r[4]
        if code and example:
            entry = f"{example} ({commodity_title})" if commodity_title else str(example)
            tech[code].append(entry)
    print(f"    {sum(len(v) for v in tech.values()):,} tech skill entries across {len(tech):,} occupations")

    rows: list[dict] = []
    for code, data in occupations.items():
        title = (data["title"] or "").strip()
        description = (data["description"] or "").strip()
        if not title or not description:
            continue

        enriched = description
        occ_tasks = tasks.get(code, [])[:_ONET_TASK_LIMIT]
        if occ_tasks:
            enriched += "\n\nCore tasks:\n" + "\n".join(f"- {t}" for t in occ_tasks)

        occ_tech = list(dict.fromkeys(tech.get(code, [])))[:_ONET_TECH_LIMIT]
        if occ_tech:
            enriched += "\n\nTechnology skills:\n" + "\n".join(f"- {t}" for t in occ_tech)

        rows.append({
            "conceptUri": f"onet:{code}",
            "preferredLabel": title,
            "altLabels": "",
            "description": enriched,
        })

    return rows


# ── Role Templates from O*NET ─────────────────────────────────────────────────

_ZONE_TO_LEVEL = {1: "entry", 2: "junior", 3: "mid", 4: "senior", 5: "expert"}
_ZONE_TO_EXP = {
    1: {"min": 0, "max": 1},
    2: {"min": 0, "max": 2},
    3: {"min": 1, "max": 4},
    4: {"min": 2, "max": 7},
    5: {"min": 5, "max": 15},
}
_EDU_LABELS = {
    1: "Less than High School",
    2: "High School Diploma",
    3: "Post-Secondary Certificate",
    4: "Some College",
    5: "Associate's Degree",
    6: "Bachelor's Degree",
    7: "Post-Baccalaureate Certificate",
    8: "Master's Degree",
    9: "Post-Master's Certificate",
    10: "First Professional Degree",
    11: "Doctoral Degree",
    12: "Post-Doctoral Training",
}

# Importance threshold on the 1–5 IM scale
_REQUIRED_THRESHOLD = 3.5
_NICE_THRESHOLD = 2.5


def _read_txt(path: str, delimiter: str = "\t") -> list[dict]:
    with open(path, encoding="utf-8-sig") as f:
        return list(csv.DictReader(f, delimiter=delimiter))


def build_role_templates_from_onet(
    onet_zip_path: str,
    txt_dir: str,
) -> list[dict]:
    """Build role_templates_real.json from O*NET 30.2 data.

    Uses the Excel zip for Occupation Data + Job Zones and the individual
    txt files (Skills, Knowledge, Technology Skills, Task Statements,
    Education Training and Experience) for skills/tasks detail.
    """
    import openpyxl  # noqa: PLC0415

    txt = Path(txt_dir)
    z = zipfile.ZipFile(onet_zip_path)

    # ── 1. Occupation titles + descriptions ───────────────────────────────────
    print("  Loading Occupation Data ...")
    occupations: dict[str, dict] = {}
    wb = openpyxl.load_workbook(
        io.BytesIO(z.read("db_30_2_excel/Occupation Data.xlsx")),
        read_only=True, data_only=True,
    )
    for row in list(wb.active.iter_rows(values_only=True))[1:]:
        code, title, desc = row[0], row[1], row[2]
        if code and title:
            occupations[str(code)] = {
                "title": str(title),
                "description": str(desc) if desc else "",
            }
    wb.close()
    print(f"    {len(occupations):,} occupations")

    # ── 2. Job Zones (experience / preparation level proxy) ───────────────────
    print("  Loading Job Zones ...")
    job_zones: dict[str, int] = {}
    wb = openpyxl.load_workbook(
        io.BytesIO(z.read("db_30_2_excel/Job Zones.xlsx")),
        read_only=True, data_only=True,
    )
    for row in list(wb.active.iter_rows(values_only=True))[1:]:
        code, _, zone = row[0], row[1], row[2]
        if code and zone:
            job_zones[str(code)] = int(zone)
    wb.close()
    z.close()

    # ── 3. Skills (importance-rated) ──────────────────────────────────────────
    print("  Loading Skills.txt ...")
    skills_req: dict[str, list] = defaultdict(list)
    skills_nice: dict[str, list] = defaultdict(list)
    for row in _read_txt(str(txt / "Skills.txt")):
        if row["Scale ID"] != "IM":
            continue
        try:
            score = float(row["Data Value"])
        except ValueError:
            continue
        code, name = row["O*NET-SOC Code"], row["Element Name"]
        if score >= _REQUIRED_THRESHOLD:
            skills_req[code].append((score, name))
        elif score >= _NICE_THRESHOLD:
            skills_nice[code].append((score, name))
    for d in (skills_req, skills_nice):
        for code in d:
            d[code] = [n for _, n in sorted(d[code], key=lambda x: -x[0])]
    # cap lists so content stays concise
    for code in skills_req:
        skills_req[code] = skills_req[code][:12]
    for code in skills_nice:
        skills_nice[code] = skills_nice[code][:6]

    # ── 4. Knowledge areas ────────────────────────────────────────────────────
    print("  Loading Knowledge.txt ...")
    knowledge: dict[str, list] = defaultdict(list)
    for row in _read_txt(str(txt / "Knowledge.txt")):
        if row["Scale ID"] != "IM":
            continue
        try:
            score = float(row["Data Value"])
        except ValueError:
            continue
        if score >= 3.0:
            knowledge[row["O*NET-SOC Code"]].append((score, row["Element Name"]))
    for code in knowledge:
        knowledge[code] = [n for _, n in sorted(knowledge[code], key=lambda x: -x[0])[:8]]

    # ── 5. Technology Skills (hot-technology first, then in-demand) ───────────
    print("  Loading Technology Skills.txt ...")
    tech: dict[str, list] = defaultdict(list)
    for row in _read_txt(str(txt / "Technology Skills.txt")):
        example = row.get("Example", "").strip()
        if not example:
            continue
        hot = row.get("Hot Technology", "N") == "Y"
        in_demand = row.get("In Demand", "N") == "Y"
        tech[row["O*NET-SOC Code"]].append((hot, in_demand, example))
    for code in tech:
        seen: set[str] = set()
        deduped: list[str] = []
        for _, _, ex in sorted(tech[code], key=lambda x: (-x[0], -x[1], x[2])):
            if ex not in seen:
                seen.add(ex)
                deduped.append(ex)
        tech[code] = deduped[:15]

    # ── 6. Core task statements ───────────────────────────────────────────────
    print("  Loading Task Statements.txt ...")
    tasks: dict[str, list] = defaultdict(list)
    for row in _read_txt(str(txt / "Task Statements.txt")):
        if row.get("Task Type", "").strip().lower() == "core":
            tasks[row["O*NET-SOC Code"]].append(row["Task"])
    for code in tasks:
        tasks[code] = tasks[code][:8]

    # ── 7. Required education level (highest-percentage category) ─────────────
    print("  Loading Education, Training, and Experience.txt ...")
    edu_data: dict[str, list] = defaultdict(list)
    for row in _read_txt(str(txt / "Education, Training, and Experience.txt")):
        if (
            row.get("Element Name") == "Required Level of Education"
            and row.get("Scale ID") == "RL"
        ):
            try:
                pct = float(row["Data Value"])
                cat = int(row["Category"])
                edu_data[row["O*NET-SOC Code"]].append((pct, cat))
            except (ValueError, KeyError):
                pass
    edu_label: dict[str, str] = {}
    for code, entries in edu_data.items():
        if entries:
            _, cat = max(entries, key=lambda x: x[0])
            edu_label[code] = _EDU_LABELS.get(cat, "")

    # ── 8. Assemble templates ─────────────────────────────────────────────────
    print("  Assembling role templates ...")
    templates: list[dict] = []
    for code, occ in occupations.items():
        zone = job_zones.get(code, 3)
        level = _ZONE_TO_LEVEL.get(zone, "mid")
        exp = _ZONE_TO_EXP.get(zone, {"min": 1, "max": 5})

        req = skills_req.get(code, [])
        nice = skills_nice.get(code, [])
        know = knowledge.get(code, [])
        tech_stack = tech.get(code, [])
        task_list = tasks.get(code, [])
        edu = edu_label.get(code, "")

        description_parts = [occ["description"]]
        if task_list:
            description_parts.append(
                "\nCore tasks:\n" + "\n".join(f"- {t}" for t in task_list)
            )
        if know:
            description_parts.append(
                "\nKnowledge areas:\n" + "\n".join(f"- {k}" for k in know)
            )
        if edu:
            description_parts.append(f"\nTypical education: {edu}")

        templates.append({
            "id": code.replace(".", "-"),
            "role": occ["title"],
            "level": level,
            "description": "\n".join(description_parts),
            "required_skills": req,
            "nice_to_have": nice,
            "experience_years": exp,
            "certifications": [],
            "region": "Global",
            "industries": [],
            "tech_stack": tech_stack,
            "education": edu,
            "onet_code": code,
            "job_zone": zone,
        })

    return templates


# ── Writer ────────────────────────────────────────────────────────────────────

_FIELDNAMES = ["conceptUri", "preferredLabel", "altLabels", "description"]


def write_csv(path: Path, rows: list[dict]) -> None:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=_FIELDNAMES)
    writer.writeheader()
    writer.writerows(rows)
    path.write_text(buf.getvalue(), encoding="utf-8")
    size_kb = path.stat().st_size // 1024
    print(f"  Wrote {len(rows):,} rows -> {path.name} ({size_kb:,} KB)")


def write_json(path: Path, data: list[dict]) -> None:
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
    size_kb = path.stat().st_size // 1024
    print(f"  Wrote {len(data):,} records -> {path.name} ({size_kb:,} KB)")


# ── CLI ───────────────────────────────────────────────────────────────────────

def main() -> None:
    _default_out = str(Path(__file__).resolve().parent.parent / "data" / "knowledge-base")

    parser = argparse.ArgumentParser(
        description="Prepare real ESCO + O*NET data for RAG ingestion.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "--esco-zip",
        default=r"C:\Users\User\Downloads\ESCO dataset - v1.2.1 - classification - en - csv.zip",
        help="Path to the ESCO v1.2.1 CSV zip",
    )
    parser.add_argument(
        "--onet-csv",
        default=r"C:\Users\User\Downloads\2019_Occupations.csv",
        help="Path to the O*NET 2019 occupations CSV (fallback if Excel zip fails)",
    )
    parser.add_argument(
        "--onet-zip",
        default=r"C:\Users\User\Downloads\db_30_2_excel.zip",
        help="Path to the O*NET 30.2 Excel zip",
    )
    parser.add_argument(
        "--onet-txt-dir",
        default=r"C:\Users\User\Downloads",
        help=(
            "Directory containing the O*NET individual txt files: Skills.txt, "
            "Knowledge.txt, Technology Skills.txt, Task Statements.txt, "
            "Education, Training, and Experience.txt"
        ),
    )
    parser.add_argument(
        "--output-dir", default=_default_out,
        help=f"Where to write output files (default: {_default_out})",
    )
    parser.add_argument(
        "--skip-esco", action="store_true", help="Skip ESCO processing",
    )
    parser.add_argument(
        "--skip-onet", action="store_true", help="Skip O*NET occupation CSV processing",
    )
    parser.add_argument(
        "--build-templates", action="store_true",
        help=(
            "Build role_templates_real.json from O*NET 30.2 data. "
            "Requires --onet-zip and --onet-txt-dir."
        ),
    )
    args = parser.parse_args()

    out = Path(args.output_dir)
    out.mkdir(parents=True, exist_ok=True)
    print(f"\nOutput directory: {out.resolve()}\n")

    total_rows = 0

    if not args.skip_esco:
        print("=== ESCO ===")
        occupations, skills, relations = _read_esco_zip(args.esco_zip)

        print("  Building enriched occupation documents ...")
        occ_enriched = build_esco_occupations_enriched(occupations, relations)
        write_csv(out / "esco_occupations_enriched.csv", occ_enriched)
        total_rows += len(occ_enriched)

        print("  Building skill documents ...")
        skill_docs = build_esco_skills(skills)
        write_csv(out / "esco_skills.csv", skill_docs)
        total_rows += len(skill_docs)
        print()

    if not args.skip_onet:
        print("=== O*NET 30.2 (Excel) — occupation CSVs ===")
        onet_enriched = build_onet_enriched_from_zip(args.onet_zip)
        if onet_enriched:
            write_csv(out / "onet_occupations_enriched.csv", onet_enriched)
            total_rows += len(onet_enriched)
        else:
            print("  Falling back to O*NET 2019 CSV ...")
            onet_rows = build_onet_from_csv(args.onet_csv)
            write_csv(out / "onet_occupations_enriched.csv", onet_rows)
            total_rows += len(onet_rows)
        print()

    if args.build_templates:
        print("=== O*NET 30.2 — role templates ===")
        _required_txt = [
            "Skills.txt",
            "Knowledge.txt",
            "Technology Skills.txt",
            "Task Statements.txt",
            "Education, Training, and Experience.txt",
        ]
        txt_dir = Path(args.onet_txt_dir)
        missing = [f for f in _required_txt if not (txt_dir / f).exists()]
        if missing:
            print(f"  ERROR: missing txt files in {txt_dir}:")
            for f in missing:
                print(f"    - {f}")
            print("  Download them from: https://www.onetcenter.org/database.html#individual-files")
        else:
            templates = build_role_templates_from_onet(args.onet_zip, args.onet_txt_dir)
            out_path = out / "role_templates_real.json"
            write_json(out_path, templates)
            total_rows += len(templates)
            print()

    print(f"Done. {total_rows:,} total documents written to {out.resolve()}")
    print()
    print("Ingest commands:")
    if not args.skip_esco:
        print(f"  esco occupations : POST /api/v1/admin/kb/ingest  {{'doc_types': ['esco']}}")
        print(f"  esco skills      : POST /api/v1/admin/kb/ingest  {{'doc_types': ['esco'], 'source_overrides': {{'esco': '{out}/esco_skills.csv'}}}}")
    if not args.skip_onet:
        print(f"  onet occupations : POST /api/v1/admin/kb/ingest  {{'doc_types': ['onet']}}")
    if args.build_templates:
        print(f"  role templates   : POST /api/v1/admin/kb/ingest  {{'doc_types': ['role_templates'], 'source_overrides': {{'role_templates': '{out}/role_templates_real.json'}}}}")


if __name__ == "__main__":
    main()
